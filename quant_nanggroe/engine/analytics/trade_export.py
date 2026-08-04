"""Export trade journal (with metacognition) to Excel and PDF.

Reads the persisted trade journal (``paper_state/trades.json``) where each
ClosedTrade carries its ``awareness`` dict (APA/KENAPA/BAGAIMANA/MENGAPA/KE MANA
+ exit cause + self-evolve lesson). Produces:
  - Excel (.xlsx) via openpyxl — one row per trade, awareness columns expanded.
  - PDF via reportlab — a per-trade report card with the full awareness narrative.

Both are honest exports: they serialise exactly what was recorded, never
recompute or invent rationale.
"""

from __future__ import annotations

import json
import os
from datetime import datetime, timezone
from typing import Any, Dict, List, Optional

# journal path mirrors api/routes/trade_history.py:_try_journal
_JOURNAL = os.path.abspath(
    os.path.join(os.path.dirname(__file__), "..", "..", "..", "paper_state", "trades.json")
)


def load_journal(limit: int = 500) -> List[Dict[str, Any]]:
    if not os.path.isfile(_JOURNAL):
        return []
    with open(_JOURNAL, encoding="utf-8") as f:
        try:
            data = json.load(f)
        except Exception:
            return []
    if isinstance(data, dict):
        data = data.get("trades", [])
    return data[-limit:]


# Flat column order for the Excel export
_COLUMNS = [
    ("trade_id", "trade_id"),
    ("symbol", "symbol"),
    ("strategy", "strategy_name"),
    ("side", "side"),
    ("action", "awareness.action"),
    ("entry_price", "entry_price"),
    ("exit_price", "exit_price"),
    ("sl", "awareness.sl"),
    ("tp", "awareness.tp"),
    ("pnl", "pnl"),
    ("outcome", "awareness.outcome"),
    ("exit_trigger", "awareness.exit_trigger"),
    ("entry_trigger", "awareness.entry_trigger"),
    ("confidence", "awareness.confidence"),
    ("regime", "awareness.regime"),
    ("KE_MANA_thesis", "awareness.target_thesis"),
    ("MENGAPA_strategy_thesis", "awareness.strategy_thesis"),
    ("BAGAIMANA_fill", "awareness.fill_note"),
    ("exit_reason", "awareness.exit_reason"),
    ("lesson", "awareness.lesson"),
    ("entry_time", "entry_time"),
    ("exit_time", "exit_time"),
]


def _dig(d: Dict[str, Any], path: str) -> Any:
    """Fetch a nested key like 'awareness.outcome' from a dict."""
    cur: Any = d
    for part in path.split("."):
        if isinstance(cur, dict) and part in cur:
            cur = cur[part]
        else:
            return ""
    return cur


def export_excel(path: str, limit: int = 500) -> str:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill

    trades = load_journal(limit)
    wb = Workbook()
    ws = wb.active
    ws.title = "Trades"

    headers = [c[0] for c in _COLUMNS]
    ws.append(headers)
    hdr_font = Font(bold=True, color="FFFFFF")
    hdr_fill = PatternFill("solid", fgColor="1F4E78")
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = Alignment(vertical="center", wrap_text=True)

    for t in trades:
        row = [_dig(t, c[1]) for c in _COLUMNS]
        ws.append(row)

    # widen + wrap the narrative columns
    widths = {14: 14, 15: 22, 16: 40, 18: 40, 19: 55, 20: 55}
    for idx, w in widths.items():
        if idx <= len(headers):
            ws.column_dimensions[ws.cell(row=1, column=idx).column_letter].width = w
    for r in range(2, len(trades) + 2):
        for c in (15, 16, 18, 19, 20):
            if c <= len(headers):
                ws.cell(row=r, column=c).alignment = Alignment(wrap_text=True, vertical="top")

    wb.save(path)
    return path


def export_pdf(path: str, limit: int = 500) -> str:
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import mm
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import (
        SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, HRFlowable,
    )

    trades = load_journal(limit)
    styles = getSampleStyleSheet()
    title = styles["Title"]
    h2 = ParagraphStyle("h2", parent=styles["Heading2"], textColor=colors.HexColor("#1F4E78"))
    body = ParagraphStyle("body", parent=styles["BodyText"], fontSize=9, leading=12)
    small = ParagraphStyle("small", parent=styles["BodyText"], fontSize=8, leading=10, textColor=colors.grey)

    doc = SimpleDocTemplate(path, pagesize=A4, topMargin=15 * mm, bottomMargin=15 * mm)
    story = [Paragraph("QNA Autonomous — Trade Awareness Report", title),
             Paragraph(f"Generated {datetime.now(timezone.utc).isoformat()} · {len(trades)} trades", small),
             Spacer(1, 6)]

    for t in trades:
        aw = t.get("awareness", {}) or {}
        sym = t.get("symbol", "?")
        strat = t.get("strategy_name", "?")
        outcome = aw.get("outcome", "")
        trig = aw.get("action", "")
        lines = [
            f"<b>TRADE</b> {t.get('trade_id','')} · {sym} · {strat} · {t.get('side','')} · {outcome}",
            f"<b>APA (what):</b> {trig} entry={t.get('entry_price','')} exit={t.get('exit_price','')} sl={aw.get('sl','')} tp={aw.get('tp','')} pnl={t.get('pnl','')}",
            f"<b>KENAPA (why in):</b> {aw.get('entry_trigger','')} · conf={aw.get('confidence','')} · confluence={aw.get('confluence','')}",
            f"<b>MENGAPA (regime/why):</b> {aw.get('regime','')} — {aw.get('regime_reason','')} · {aw.get('strategy_thesis','')}",
            f"<b>KE MANA (intent):</b> {aw.get('target_thesis','')} · RR~{aw.get('expected_rr','')} · {aw.get('holding_intent','')}",
            f"<b>BAGAIMANA (how):</b> venue={aw.get('execution_venue','')} · {aw.get('fill_note','')}",
            f"<b>EXIT:</b> {aw.get('exit_trigger','')} — {aw.get('exit_reason','')}",
            f"<b>LESSON (self-evolve):</b> {aw.get('lesson','')}",
        ]
        for ln in lines:
            story.append(Paragraph(ln, body))
        story.append(HRFlowable(width="100%", thickness=0.5, color=colors.lightgrey))
        story.append(Spacer(1, 4))

    doc.build(story)
    return path


def export(format: str = "excel", limit: int = 500) -> str:
    """Return a path to the generated file. format in {excel, pdf}."""
    out_dir = os.path.dirname(_JOURNAL)
    os.makedirs(out_dir, exist_ok=True)
    if format == "pdf":
        return export_pdf(os.path.join(out_dir, "qna_trade_awareness.pdf"), limit)
    return export_excel(os.path.join(out_dir, "qna_trade_awareness.xlsx"), limit)
